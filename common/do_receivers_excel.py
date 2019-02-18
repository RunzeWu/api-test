# ！/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time     :2019/1/12 20:19
# @Author   :Yosef
# E-mail    :wurz529@foxmail.com
# File      :do_receivers_excel.py
# Software  :PyCharm Community Edition
import openpyxl
from common.config import ReadConfig
from common.mylog import Mylog
from common import contants

mylog = Mylog("读取excel文件")


class TestCase:
    def __init__(self):
        self.id = None
        self.name = None
        self.email = None


class DoExcel:
    def __init__(self, filepath, sheet_name):
        self.filepath = filepath
        self.sheet_name = sheet_name
        self.conf = ReadConfig().get_value("receiver_id", "button")

    def read_data(self):
        wb = openpyxl.load_workbook(self.filepath)
        sh = wb[self.sheet_name]
        testdata = []

        if self.conf == "all":
            for i in range(2, sh.max_row + 1):
                item = TestCase()
                item.id = sh.cell(i, 1).value
                item.name = sh.cell(i, 2).value
                item.email = sh.cell(i, 3).value
                testdata.append(item)
        else:
            try:
                conf = eval(self.conf)
            except Exception:
                mylog.error("请检查{}文件中[receiver_id]的button的value是否符合要求！！".format(contants.receiver_file))
                raise Exception("请检查{}文件中[receiver_id]的button的value是否符合要求！！".format(contants.receiver_file))
            new_conf = []
            for run_case_id in conf:
                run_case_row = run_case_id + 1
                new_conf.append(run_case_row)
            for i in new_conf:
                item = TestCase()
                item.id = sh.cell(i, 1).value
                item.name = sh.cell(i, 2).value
                item.email = sh.cell(i, 3).value
                testdata.append(item)

        receivers = []
        for item in testdata:
            receivers.append(item.email)

        return receivers


if __name__ == '__main__':
    excel = DoExcel(contants.receiver_file, "receivers").read_data()
    print(excel)
